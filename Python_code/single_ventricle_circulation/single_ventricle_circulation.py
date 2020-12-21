import sys
import os
import numpy as np
import pandas as pd
# import cProfile
# import timeit
# from openpyxl import Workbook
# from scipy import signal
# from scipy.integrate import solve_ivp
# from scipy.constants import mmHg as mmHg_in_pascals


from .half_sarcomere import half_sarcomere as hs

from protocol import protocol as prot

# from modules.SystemControl import system_control as syscon
# from modules.Perturbation import perturbation as pert
# from modules.Growth import growth as gr

import json

class single_ventricle_circulation():

    """Class for a single ventricle circulation"""
    from .implement import implement_time_step, update_data_holders,analyze_data
    from .display import display_simulation, display_flows, display_pv_loop

    def __init__(self, model_json_file_string):

        from .implement import return_lv_circumference,return_lv_pressure

        # Check for file
        if (model_json_file_string==[]):
            print('No model file specified. Cannot create model')
            return
        
        # Load the model as a dict
        with open(model_json_file_string,'r') as f:
            sc_struct = json.load(f)

        # Initialize the circulation object using data from the json file
        circ_struct = sc_struct["circulation"]

        # Define vessel list
        vessels_list = ['aorta','arteries','arterioles','capillaries',
                        'venules','veins']

        self.circ_model = dict()
        self.circ_model['no_of_compartments'] = int(circ_struct['no_of_compartments'])
        self.circ_model['blood_volume'] = circ_struct['blood_volume']
        self.circ_model['body_surface_area'] = circ_struct['body_surface_area']
        
        for v in vessels_list:
            for t in ['resistance','compliance']:
                n = ('%s_%s') % (v,t)
                self.circ_model[n]=circ_struct[v][t]
                
        for t in ['inter_beat_interval','resistance','wall_volume',
                  'slack_volume','wall_density']:
            self.circ_model[('%s_%s') % ('ventricle',t)] = \
                circ_struct['ventricle'][t]

        self.circ_model['lv_mass'] = self.circ_model['ventricle_wall_volume'] * \
            self.circ_model['ventricle_wall_density']
            
        # Build the compliance and resistance arrays
        self.circ_model['compliance'] = []
        for v in vessels_list:
            c = self.circ_model[('%s_compliance' % v)]
            self.circ_model['compliance'].append(c)
        # Add in 0 for ventricular compliance
        self.circ_model['compliance'].append(0)
        # Convert to numpy array
        self.circ_model['compliance'] = np.array(self.circ_model['compliance'])
        
        self.circ_model['resistance'] = []
        v_list = vessels_list
        v_list.append('ventricle')
        for v in v_list:
            r = self.circ_model[('%s_resistance' % v)]
            self.circ_model['resistance'].append(r)
        self.circ_model['resistance'] = np.array(self.circ_model['resistance'])

        # Pull off the half_sarcomere parameters and make a half-sarcomere
        hs_params = sc_struct["half_sarcomere"]
        self.hs = hs.half_sarcomere(hs_params)

        # Deduce the hsl where force is zero and set the hsl to that length
        self.slack_hsl = self.hs.myof.return_hs_length_for_force(0.0)
        self.delta_hsl = self.slack_hsl - self.hs.hs_length
        self.hs.update_simulation(0.0,self.delta_hsl, 0.0)

        # Update some stuff that will change with growth and perturbations
        self.ventricle_wall_volume = self.circ_model['ventricle_wall_volume']
        self.ventricle_slack_volume = self.circ_model['ventricle_slack_volume']
        self.blood_volume = self.circ_model['blood_volume']

        # Deduce the slack circumference of the ventricle and set that
        self.lv_circumference =\
            return_lv_circumference(self, self.ventricle_slack_volume)
        self.n_hs = 10e9*self.lv_circumference / self.slack_hsl

        internal_r = np.power((3.0 * 0.001 * 1.5*self.ventricle_slack_volume)/
                    (2.0 * np.pi), (1.0 / 3.0))

        internal_area = 2.0 * np.pi * np.power(internal_r, 2.0)
        self.wall_thickness = 0.001 * self.ventricle_wall_volume /internal_area

        # if "growth" in single_circulation_simulation:

        #     from modules.Growth import growth as gr

        #     growth_params = single_circulation_simulation["growth"]
        #     start_index = int(growth_params["start_index"][0])

        #     self.driven_signal = growth_params["driven_signal"][0]

        #     if self.driven_signal != "stress" and self.driven_signal!="ATPase":
        #         print('Growth driven signal is not defined correctly!')

        #     initial_numbers_of_hs = self.n_hs
        #     self.gr = \
        #     gr.growth(growth_params,initial_numbers_of_hs,self.hs,circ_params,
        #                     self.output_buffer_size)

        #     self.growth_activation_array[start_index:] = True
        #     self.growth_activation = self.growth_activation_array[0]

        # # Baro
        # self.syscon=syscon.system_control(self.sys_params,hs_params,self.hs,
        #                 circ_params,self.output_buffer_size)
        # self.baro_activation_array = np.full(self.output_buffer_size+1,False)
        # if "baroreceptor" in self.sys_params:

        #     start_index = int(self.sys_params["baroreceptor"]["start_index"][0])
        #     self.baro_activation_array[start_index:]=True
        # self.baro_activation = self.baro_activation_array[0]

        print("hsl: %f" % self.hs.hs_length)
        print("slack hsl: %f" % self.slack_hsl)
        print("slack_lv_circumference %f" % self.lv_circumference)

        # Set the initial volumes with most of the blood in the veins
        initial_ventricular_volume = 1.5 * self.ventricle_slack_volume
        self.v = np.zeros(self.circ_model['no_of_compartments'])
        self.v[-2] = self.blood_volume - initial_ventricular_volume
        self.v[-1] = initial_ventricular_volume

        # Deduce the pressures
        self.p = np.zeros(self.circ_model['no_of_compartments'])
        for i in np.arange(0, self.circ_model['no_of_compartments']-1):
            self.p[i] = self.v[i] / self.circ_model['compliance'][i]
        self.p[-1] = return_lv_pressure(self,self.v[-1])



#         # Store the first values
#         self.data.at[0, 'pressure_aorta'] = self.p[0]
#         self.data.at[0, 'pressure_arteries'] = self.p[1]
#         self.data.at[0, 'pressure_arterioles'] = self.p[2]
#         self.data.at[0, 'pressure_capillaries'] = self.p[3]
#         self.data.at[0, 'pressure_veins'] = self.p[4]
#         self.data.at[0, 'pressure_ventricle'] = self.p[5]

#         self.data.at[0, 'volume_aorta'] = self.v[0]
#         self.data.at[0, 'volume_arteries'] = self.v[1]
#         self.data.at[0, 'volume_arterioles'] = self.v[2]
#         self.data.at[0, 'volume_capillaries'] = self.v[3]
#         self.data.at[0, 'volume_veins'] = self.v[4]
#         self.data.at[0, 'volume_ventricle'] = self.v[-1]

#         self.data.at[0, 'volume_aortic_regurgitation'] = self.vl[0]
#         self.data.at[0, 'volume_mitral_regurgitation'] = self.vl[1]

#         """if self.hs.ATPase_activation:
#             self.ATPase = return_ATPase(self)
#             self.data.at[0, 'ATPase'] = self.ATPase"""

#         self.prof_activation = \
#             single_circulation_simulation["profiling"]["profiling_activation"][0]

    def create_data_structure(self):
        
        data_fields = ['time',
                  'pressure_aorta','pressure_arteries','pressure_arterioles',
                  'pressure_capillaries','pressure_venules','pressure_veins',
                  'presure_ventricle',
                  'volume_aorta','volume_arteries','volume_arterioles',
                  'volume_capillaries','volume_venules','volume_veins',
                  'volume_ventricle',
                  'ventricle_wall_thickness', 'ventricle_wall_volume',
                  'aorta_resistance','arteries_resistance','arterioles_resistance',
                  'capillaries_resistance','venules_resistance','veins_resistance',
                  'ventricle_resistance',
                  'aorta_compliance','arteries_compliance','arterioles_compliance',
                  'capillaries_compliance','venules_compliance','veins_compliance',
                  'ventricle_resistance',
                  'flow_ventricle_to_aorta', 'flow_aorta_to_arteries',
                  'flow_arteries_to_arterioles', 'flow_arterioles_to_capillaries',
                  'flow_capillaries_to_venules', 'flow_venules_to_veins',
                  'flow_veins_to_ventricle',
                  'flow_ventricle_to_veins', 'flow_veins_to_venules',
                  'flow_venules_to_capillaries', 'flow_capillaries_to_arterioles',
                  'flow_arterioles_to_arteries', 'flow_arteries_to_aorta',
                  'flow_aorta_to_ventricle']
        
        self.data = pd.DataFrame()
        z = np.zeros(self.prot.data['no_of_time_points'])
        for f in data_fields:
            s = pd.Series(data=z, name=f)
            self.data = pd.concat([self.data, s], axis=1)
        for f in self.hs.data_fields:
            s = pd.Series(data=z, name=f)
            self.data = pd.concat([self.data, s], axis=1)
        for f in self.hs.myof.data_fields:
            s = pd.Series(data=z, name=f)
            self.data = pd.concat([self.data, s], axis=1)
            
        print(self.data)
       

    def run_simulation(self, protocol_file_string=[], output_structure_file_string=[]):
        # Run the simulation

        # Load the protocol
        if (protocol_file_string==[]):
            print("No protocol_file_string. Exiting")
            return
        self.prot = prot.protocol(protocol_file_string)
        
        # Create the data structure
        self.create_data_structure()


        
        # Load the output_structure
        if (output_structure_file_string==[]):
            print("No output_structure_file_string. Exiting")
            return
        
        with open(output_structure_file_string,'r') as osf:
            output_struct = json.load(osf)
            print(output_struct)
        
        return
        



        # Set up some values for the simulation
        no_of_time_points = \
            int(self.sys_params["simulation"]["no_of_time_points"][0])

        activation_duty_ratio = \
            float(self.sys_params["simulation"]["duty_ratio"][0])

        t = self.dt*np.arange(1, no_of_time_points+1)

        # Apply profiling befor running the simulation
        if self.prof_activation:
            pr = cProfile.Profile()
            pr.enable()

        # Run the simulation
        for i in np.arange(np.size(t)):

            if self.pert_activation:
                # Apply volume perturbation to veins
                self.v[-2] = self.v[-2] + self.volume_perturbation[i]
                # Apply valve perturbation
                    #aortic
                self.aortic_valve_perturbation_factor = \
                self.aortic_valve_perturbation[i]
                    #mitral
                self.mitral_valve_perturbation_factor = \
                self.mitral_valve_perturbation[i]
                # Apply perturbation to compliances
                self.compliance[0]=self.compliance[0]+\
                            self.aorta_compliance_perturbation[i]
                self.compliance[2]=self.compliance[2] +\
                            self.capillaries_compliance_perturbation[i]
                self.compliance[-2]=self.compliance[-2] +\
                            self.venous_compliance_perturbation[i]
                # Apply perturbation to resistances
                self.resistance[0]=self.resistance[0]+\
                            self.aorta_resistance_perturbation[i]
                self.resistance[2]=self.resistance[2] +\
                            self.capillaries_resistance_perturbation[i]
                self.resistance[-2]=self.resistance[-2] +\
                            self.venous_resistance_perturbation[i]
                self.resistance[-1] = self.resistance[-1] +\
                            self.ventricle_resistance_perturbation[i]
                # Apply perturbation to myosim
                self.hs.myof.k_1 = self.hs.myof.k_1 + self.k_1_perturbation[i]
                self.hs.myof.k_2 = self.hs.myof.k_2 + self.k_2_perturbation[i]
                self.hs.myof.k_4_0 = self.hs.myof.k_4_0 +self.k_4_0_perturbation[i]

                self.hs.membr.Ca_Vmax_up_factor =\
                            self.hs.membr.Ca_Vmax_up_factor + self.ca_uptake_perturbation[i]
                self.hs.membr.Ca_V_leak_factor =\
                            self.hs.membr.Ca_V_leak_factor + self.ca_leak_perturbation[i]
                self.hs.membr.g_CaL_factor =  \
                            self.hs.membr.g_CaL_factor + self.g_cal_perturbation[i]
            # Apply growth activation
            self.growth_activation = self.growth_activation_array[i]
            # Apply baro activation
            self.baro_activation = self.baro_activation_array[i]
            # Apply heart rate activation
            activation_level=self.syscon.return_activation()

            # Display
            if ( (i % 2000) == 0):
                print("Blood volume: %.3g, %.0f %% complete" %
                      (np.sum(self.v), (100*i/np.size(t))))

            implement_time_step(self, self.dt, activation_level,i)
            update_data_holders(self, self.dt, activation_level)

        # Concatenate data structures
        self.data = pd.concat([self.data, self.hs.hs_data, self.syscon.sys_data],axis=1)

        if self.growth_activation:
            self.data =pd.concat([self.data,self.gr.gr_data],axis=1)

        # Get output for multithreading
        #if self.multithreading_activation:
        #    return self.data




        if self.prof_activation:
            pr.disable()
            pr.print_stats()
        # Make plots
        # Circulation
        display_simulation(self.data,
                           self.output_parameters["summary_figure"][0],dpi=300)#,[75,120])#,[81.6,82.6])

        display_flows(self.data,
                      self.output_parameters["flows_figure"][0],dpi=300)
        display_pv_loop(self.data,
                        self.output_parameters["pv_figure"][0],dpi=300)#,[[78.5,79.8],[142.8,143.8]]

        if self.baro_activation:
            syscon.system_control.display_baro_results(self.data,
                            self.output_parameters["baro_figure"][0],dpi=300)
            syscon.system_control.display_arterial_pressure(self.data,
                            self.output_parameters["circulatory"][0],dpi=300)

        # Half-sarcomere
        hs.half_sarcomere.display_fluxes(self.data,
                               self.output_parameters["hs_fluxes_figure"][0],dpi=300)#,[30,60])

        #Growth
        if self.growth_activation:

            gr.growth.display_growth(self.data,
            self.output_parameters["growth_figure"][0],self.driven_signal)

            gr.growth.display_growth_summary(self.data,
            self.output_parameters["growth_summary"][0],self.driven_signal)


        if self.hs.ATPase_activation:
            gr.growth.display_ATPase(self.data,self.output_parameters["ATPase"][0])
#        display_regurgitation(self.data,
#                self.output_parameters["regurg_fig"][0])
        if self.saving_data_activation:

            print("Data is saving to %s format!"%self.output_data_format)

            data_to_be_saved = \
                self.data.loc[self.save_data_start_index:self.save_data_stop_index,:]
            if self.output_data_format == "csv":
                start_csv = timeit.default_timer()
                data_to_be_saved.to_csv(self.output_parameters['csv_file'][0])
                stop_csv = timeit.default_timer()
                csv_time = stop_csv-start_csv
                print('dumping data to .csv format took %f seconds'%csv_time)

            elif self.output_data_format == "excel":
                start_excel = timeit.default_timer()
                append_df_to_excel(self.output_parameters['excel_file'][0],data_to_be_saved,
                            sheet_name='Data',startrow=0)
                stop_excel = timeit.default_timer()
                excel_time = stop_excel-start_excel
                print('dumping data to .excel format took %f seconds'%excel_time)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
