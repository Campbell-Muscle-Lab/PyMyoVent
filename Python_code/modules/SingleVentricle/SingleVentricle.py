import sys
import os
import numpy as np
import pandas as pd
import cProfile
import timeit
from openpyxl import Workbook
from scipy import signal
from scipy.integrate import solve_ivp
from scipy.constants import mmHg as mmHg_in_pascals


from modules.MyoSim.half_sarcomere import half_sarcomere as hs
from modules.SystemControl import system_control as syscon
from modules.Perturbation import perturbation as pert
from modules.Growth import growth as gr

class single_circulation():
    """Class for a single ventricle circulation"""
    from .implement import implement_time_step, update_data_holders,analyze_data
    from .display import display_simulation, display_flows, display_pv_loop

    def __init__(self, single_circulation_simulation, xml_file_string=None):

        from .implement import return_lv_circumference,return_lv_pressure, return_ATPase
        # Pull off stuff

        #self.multithreading_activation = \
        #    single_circulation_simulation["multi_threads"]["multithreading_activation"][0]

        self.output_parameters = \
            single_circulation_simulation["output_parameters"]


        self.sys_params = single_circulation_simulation["system_control"]

        self.output_buffer_size = int(self.sys_params["simulation"]["no_of_time_points"][0])
        self.dt= float(self.sys_params["simulation"]["time_step"][0])
        self.T= float(self.sys_params["simulation"]["basal_heart_period"][0])

        # Initialize circulation object using data from the sim_object
        circ_params = single_circulation_simulation["circulation"]

        self.no_of_compartments = int(circ_params["no_of_compartments"][0])
        self.blood_volume = float(circ_params["blood"]["volume"][0])

        self.aorta_resistance = float(circ_params["aorta"]["resistance"][0])
        self.aorta_compliance = float(circ_params["aorta"]["compliance"][0])

        self.arteries_resistance = float(circ_params["arteries"]["resistance"][0])
        self.arteries_compliance = float(circ_params["arteries"]["compliance"][0])

        self.arterioles_resistance = float(circ_params["arterioles"]["resistance"][0])
        self.arterioles_compliance = float(circ_params["arterioles"]["compliance"][0])

        self.capillaries_resistance = float(circ_params["capillaries"]["resistance"][0])
        self.capillaries_compliance = float(circ_params["capillaries"]["compliance"][0])

        self.veins_resistance = float(circ_params["veins"]["resistance"][0])
        self.veins_compliance = float(circ_params["veins"]["compliance"][0])

        self.ventricle_resistance = \
            float(circ_params["ventricle"]["resistance"][0])
        self.ventricle_wall_volume = \
            float(circ_params["ventricle"]["wall_volume"][0])
        self.ventricle_wall_density = \
            float(circ_params["ventricle"]["wall_density"][0])
        self.ventricle_slack_volume = \
            float(circ_params["ventricle"]["slack_volume"][0])
        self.body_surface_area = \
            float(circ_params["ventricle"]["body_surface_area"][0])

        self.lv_mass = \
                    self.ventricle_wall_volume*self.ventricle_wall_density
        self.lv_mass_indexed = \
                    self.lv_mass/self.body_surface_area

        # Initialise the resistance and compliance arrays for calcuations
        self.resistance = np.array([self.aorta_resistance,
                                    self.arteries_resistance,
                                    self.arterioles_resistance,
                                    self.capillaries_resistance,
                                    self.veins_resistance,
                                    self.ventricle_resistance])
        self.compliance = np.array([self.aorta_compliance,
                                    self.arteries_compliance,
                                    self.arterioles_compliance,
                                    self.capillaries_compliance,
                                    self.veins_compliance,
                                    0])

        # Look for perturbations
        self.pert_activation = False
        if "perturbations" in single_circulation_simulation:

            self.pert_activation = True
            pert_params = single_circulation_simulation['perturbations']
            self.pert = pert.perturbation(pert_params,self.output_buffer_size)
            self.volume_perturbation = self.pert.volume_perturbation

            self.aortic_valve_perturbation =\
            self.pert.aortic_valve_perturbation

            self.mitral_valve_perturbation =\
            self.pert.mitral_valve_perturbation

            self.aorta_compliance_perturbation = \
            self.pert.aorta_compliance_perturbation

            self.capillaries_compliance_perturbation=\
            self.pert.capillaries_compliance_perturbation

            self.venous_compliance_perturbation =\
            self.pert.venous_compliance_perturbation

            self.aorta_resistance_perturbation=\
            self.pert.aorta_resistance_perturbation

            self.capillaries_resistance_perturbation=\
            self.pert.capillaries_resistance_perturbation

            self.venous_resistance_perturbation =\
            self.pert.venous_resistance_perturbation

            self.ventricle_resistance_perturbation =\
            self.pert.ventricle_resistance_perturbation

            self.k_1_perturbation = self.pert.k_1_perturbation

            self.k_2_perturbation = self.pert.k_2_perturbation

            self.k_4_0_perturbation = self.pert.k_4_0_perturbation

            self.ca_uptake_perturbation = self.pert.ca_uptake_perturbation

            self.ca_leak_perturbation = self.pert.ca_leak_perturbation

            self.g_cal_perturbation = self.pert.g_cal_perturbation

        # Pull off the half_sarcomere parameters
        hs_params = single_circulation_simulation["half_sarcomere"]
        self.hs = hs.half_sarcomere(hs_params, self.output_buffer_size)

        # Deduce the hsl where force is zero and set the hsl to that length
        self.slack_hsl = self.hs.myof.return_hs_length_for_force(0.0)
        self.delta_hsl = self.slack_hsl - self.hs.hs_length
        self.hs.update_simulation(0.0,self.delta_hsl, 0.0)

        # Deduce the slack circumference of the ventricle and set that

        self.lv_circumference =\
         return_lv_circumference(self,self.ventricle_slack_volume)
        self.n_hs = 10e9*self.lv_circumference / self.slack_hsl

        internal_r = np.power((3.0 * 0.001 * 1.5*self.ventricle_slack_volume)/
                    (2.0 * np.pi), (1.0 / 3.0))

        internal_area = 2.0 * np.pi * np.power(internal_r, 2.0)
        self.wall_thickness = 0.001 * self.ventricle_wall_volume /internal_area

        # Baro

        self.syscon=syscon.system_control(self.sys_params,hs_params,self.hs,
                        circ_params,self.output_buffer_size)
        self.baro_activation_array = np.full(self.output_buffer_size+1,False)
        if "baroreceptor" in self.sys_params:
            self.membrane_kinetic_scheme = hs_params['membranes']['kinetic_scheme'][0]
            start_index = int(self.sys_params["baroreceptor"]["start_index"][0])
            self.baro_activation_array[start_index:]=True
        self.baro_activation = self.baro_activation_array[0]

        # Look for growth module
        self.growth_activation_array = np.full(self.output_buffer_size+1,False)
        self.growth_activation = self.growth_activation_array[0]

        if "growth" in single_circulation_simulation:

            from modules.Growth import growth as gr

            growth_params = single_circulation_simulation["growth"]
            start_index = int(growth_params["start_index"][0])

            self.driven_signal = growth_params["driven_signal"][0]

            if self.driven_signal != "stress" and self.driven_signal!="ATPase":
                print('Growth driven signal is not defined correctly!')

            initial_numbers_of_hs = self.n_hs
            self.gr = \
            gr.growth(growth_params,initial_numbers_of_hs,self.hs,self.syscon,circ_params,
                            self.output_buffer_size)

            self.growth_activation_array[start_index:] = True
            self.growth_activation = self.growth_activation_array[0]



        print("hsl: %f" % self.hs.hs_length)
        print("slack hsl: %f" % self.slack_hsl)
        print("slack_lv_circumference %f" % self.lv_circumference)

        # Set the initial volumes with most of the blood in the veins
        initial_ventricular_volume = 1.5 * self.ventricle_slack_volume
        self.v = np.zeros(self.no_of_compartments)
        self.v[4] = self.blood_volume - initial_ventricular_volume
        self.v[-1] = initial_ventricular_volume

        # Deduce the pressures
        self.p = np.zeros(self.no_of_compartments)
        for i in np.arange(0, self.no_of_compartments-1):
            self.p[i] = self.v[i] / self.compliance[i]
        self.p[-1] = return_lv_pressure(self,self.v[-1])

        #Valve leakages
        self.vl = np.zeros(2)

        # ATPase
#        self.ATPase_activation = \
#            single_circulation_simulation["ATPase"][0]

        # saving data
#        self.saving_data_activation = False
        self.saving_data_activation =  \
            single_circulation_simulation["saving_to_spreadsheet"]["saving_data_activation"][0]
        if self.saving_data_activation:
            self.output_data_format = \
                single_circulation_simulation["saving_to_spreadsheet"]["output_data_format"][0]
            self.save_data_start_index = \
                single_circulation_simulation["saving_to_spreadsheet"]["start_index"][0]
            self.save_data_stop_index= \
                single_circulation_simulation["saving_to_spreadsheet"]["stop_index"][0]
        # Create a pandas data structure to store data
        self.sim_time = 0.0
        self.data_buffer_index = 0
        self.data = pd.DataFrame({'time': np.zeros(self.output_buffer_size),
                                  'pressure_aorta':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_arteries':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_arterioles':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_capillaries':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_veins':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_ventricle':
                                      np.zeros(self.output_buffer_size),
                                  'volume_aorta':
                                      np.zeros(self.output_buffer_size),
                                  'volume_arteries':
                                      np.zeros(self.output_buffer_size),
                                  'volume_arterioles':
                                      np.zeros(self.output_buffer_size),
                                  'volume_capillaries':
                                      np.zeros(self.output_buffer_size),
                                  'volume_veins':
                                      np.zeros(self.output_buffer_size),
                                  'volume_ventricle':
                                      np.zeros(self.output_buffer_size),
                                  'volume_aortic_regurgitation':
                                      np.zeros(self.output_buffer_size),
                                  'volume_mitral_regurgitation':
                                      np.zeros(self.output_buffer_size),
                                  'flow_ventricle_to_aorta':
                                      np.zeros(self.output_buffer_size),
                                  'flow_aorta_to_arteries':
                                      np.zeros(self.output_buffer_size),
                                  'flow_arteries_to_arterioles':
                                      np.zeros(self.output_buffer_size),
                                  'flow_arterioles_to_capillaries':
                                      np.zeros(self.output_buffer_size),
                                  'flow_capillaries_to_veins':
                                      np.zeros(self.output_buffer_size),
                                  'flow_veins_to_ventricle':
                                      np.zeros(self.output_buffer_size),
                                  'volume_perturbation':
                                      np.zeros(self.output_buffer_size),
                                  'ventricle_wall_thickness':
                                      np.full(self.output_buffer_size, self.wall_thickness),
#                                  'ATPase':
#                                     np.zeros(self.output_buffer_size),
#                                  'ventricle_wall_volume':
#                                    np.full(self.output_buffer_size,self.ventricle_wall_volume),
                                    'ventricle_wall_mass':
                                    np.full(self.output_buffer_size,self.lv_mass),
                                    'ventricle_wall_mass_i':
                                    np.full(self.output_buffer_size,self.lv_mass_indexed),
                                    'aorta_resistance':
                                    np.full(self.output_buffer_size,self.aorta_resistance),
                                    'arteries_resistance':
                                    np.full(self.output_buffer_size,self.arteries_resistance),
                                    'arterioles_resistance':
                                    np.full(self.output_buffer_size,self.arterioles_resistance),
                                    'capillaries_resistance':
                                    np.full(self.output_buffer_size,self.capillaries_resistance),
                                    'veins_resistance':
                                    np.full(self.output_buffer_size,self.veins_resistance),
                                    'ventricle_resistance':
                                    np.full(self.output_buffer_size,self.ventricle_resistance),
                                    'aorta_compliance':
                                    np.full(self.output_buffer_size,self.aorta_compliance),
                                    'arteries_compliance':
                                    np.full(self.output_buffer_size,self.arteries_compliance),
                                    'arterioles_compliance':
                                    np.full(self.output_buffer_size,self.arterioles_compliance),
                                    'capillaries_compliance':
                                    np.full(self.output_buffer_size,self.capillaries_compliance),
                                    'veins_compliance':
                                    np.full(self.output_buffer_size,self.veins_compliance)})


        # Store the first values
        self.data.at[0, 'pressure_aorta'] = self.p[0]
        self.data.at[0, 'pressure_arteries'] = self.p[1]
        self.data.at[0, 'pressure_arterioles'] = self.p[2]
        self.data.at[0, 'pressure_capillaries'] = self.p[3]
        self.data.at[0, 'pressure_veins'] = self.p[4]
        self.data.at[0, 'pressure_ventricle'] = self.p[5]

        self.data.at[0, 'volume_aorta'] = self.v[0]
        self.data.at[0, 'volume_arteries'] = self.v[1]
        self.data.at[0, 'volume_arterioles'] = self.v[2]
        self.data.at[0, 'volume_capillaries'] = self.v[3]
        self.data.at[0, 'volume_veins'] = self.v[4]
        self.data.at[0, 'volume_ventricle'] = self.v[-1]

        self.data.at[0, 'volume_aortic_regurgitation'] = self.vl[0]
        self.data.at[0, 'volume_mitral_regurgitation'] = self.vl[1]

        """if self.hs.ATPase_activation:
            self.ATPase = return_ATPase(self)
            self.data.at[0, 'ATPase'] = self.ATPase"""

        self.prof_activation = \
            single_circulation_simulation["profiling"]["profiling_activation"][0]

    def run_simulation(self):
        # Run the simulation
        from .implement import implement_time_step, update_data_holders,analyze_data
        from .display import display_simulation, display_flows, display_pv_loop,display_N_overlap
        from .display import display_Ca,display_pres,display_simulation_publish,display_active_force,display_activation_pulse

        # Set up some values for the simulation
        no_of_time_points = \
            int(self.sys_params["simulation"]["no_of_time_points"][0])

        activation_duty_ratio = \
            float(self.sys_params["simulation"]["duty_ratio"][0])

        t = self.dt*np.arange(1, no_of_time_points+1)

        # Apply profiling befor running the simulation
        if self.prof_activation:
            pr = cProfile.Profile()
            pr.enable()

        # Run the simulation
        for i in np.arange(np.size(t)):

            if self.pert_activation:
                # Apply volume perturbation to veins
                self.v[-2] = self.v[-2] + self.volume_perturbation[i]
                # Apply valve perturbation
                    #aortic
                self.aortic_valve_perturbation_factor = \
                self.aortic_valve_perturbation[i]
                    #mitral
                self.mitral_valve_perturbation_factor = \
                self.mitral_valve_perturbation[i]
                # Apply perturbation to compliances
                self.compliance[0]=self.compliance[0]+\
                            self.aorta_compliance_perturbation[i]
                self.compliance[2]=self.compliance[2] +\
                            self.capillaries_compliance_perturbation[i]
                self.compliance[-2]=self.compliance[-2] +\
                            self.venous_compliance_perturbation[i]
                # Apply perturbation to resistances
                self.resistance[0]=self.resistance[0]+\
                            self.aorta_resistance_perturbation[i]
                self.resistance[2]=self.resistance[2] +\
                            self.capillaries_resistance_perturbation[i]
                self.resistance[-2]=self.resistance[-2] +\
                            self.venous_resistance_perturbation[i]
                self.resistance[-1] = self.resistance[-1] +\
                            self.ventricle_resistance_perturbation[i]
                # Apply perturbation to myosim
                self.hs.myof.k_1 = self.hs.myof.k_1 + self.k_1_perturbation[i]
                self.hs.myof.k_2 = self.hs.myof.k_2 + self.k_2_perturbation[i]
                self.hs.myof.k_4_0 = self.hs.myof.k_4_0 +self.k_4_0_perturbation[i]

                self.hs.membr.Ca_Vmax_up_factor =\
                            self.hs.membr.Ca_Vmax_up_factor + self.ca_uptake_perturbation[i]
                self.hs.membr.Ca_V_leak_factor =\
                            self.hs.membr.Ca_V_leak_factor + self.ca_leak_perturbation[i]
                self.hs.membr.g_CaL_factor =  \
                            self.hs.membr.g_CaL_factor + self.g_cal_perturbation[i]
            # Apply growth activation
            self.growth_activation = self.growth_activation_array[i]
            # Apply baro activation
            self.baro_activation = self.baro_activation_array[i]
            # Apply heart rate activation
            activation_level=self.syscon.return_activation()

            # Display
            if ( (i % 2000) == 0):
                print("Blood volume: %.3g, %.0f %% complete" %
                      (np.sum(self.v), (100*i/np.size(t))))

            implement_time_step(self, self.dt, activation_level,i)
            update_data_holders(self, self.dt, activation_level)

        # Concatenate data structures
        self.data = pd.concat([self.data, self.hs.hs_data, self.syscon.sys_data],axis=1)

        if self.growth_activation:
            self.data =pd.concat([self.data,self.gr.gr_data],axis=1)

        # Get output for multithreading
        #if self.multithreading_activation:
        #    return self.data




        if self.prof_activation:
            pr.disable()
            pr.print_stats()
        # Make plots
        # Circulation
        display_simulation(self.data,
                           self.output_parameters["summary_figure"][0],dpi=300)#,[75,120])#,[81.6,82.6])

        display_flows(self.data,
                      self.output_parameters["flows_figure"][0],dpi=300)
        display_pv_loop(self.data,
                        self.output_parameters["pv_figure"][0],dpi=300)#,[[78.5,79.8],[142.8,143.8]]

        if self.baro_activation:
            if self.membrane_kinetic_scheme == 'Ten_Tusscher_2004':
                syscon.system_control.display_baro_results_tt(self.data,
                                self.output_parameters["baro_figure"][0],dpi=300)
            elif self.membrane_kinetic_scheme == 'simple_2_compartment':
                syscon.system_control.display_baro_results(self.data,
                                self.output_parameters["baro_figure"][0],dpi=300)
            syscon.system_control.display_arterial_pressure(self.data,
                            self.output_parameters["circulatory"][0],dpi=300)

        # Half-sarcomere
        hs.half_sarcomere.display_fluxes(self.data,
                               self.output_parameters["hs_fluxes_figure"][0],dpi=300)#,[30,60])

        #Growth
        if self.growth_activation:

            gr.growth.display_growth(self.data,
            self.output_parameters["growth_figure"][0],self.driven_signal)

            gr.growth.display_growth_summary(self.data,
            self.output_parameters["growth_summary"][0],self.driven_signal)


        if self.hs.ATPase_activation:
            gr.growth.display_ATPase(self.data,self.output_parameters["ATPase"][0])
#        display_regurgitation(self.data,
#                self.output_parameters["regurg_fig"][0])
        if self.saving_data_activation:

            print("Data is saving to %s format!"%self.output_data_format)

            data_to_be_saved = \
                self.data.loc[self.save_data_start_index:self.save_data_stop_index,:]
            if self.output_data_format == "csv":
                start_csv = timeit.default_timer()
                data_to_be_saved.to_csv(self.output_parameters['csv_file'][0])
                stop_csv = timeit.default_timer()
                csv_time = stop_csv-start_csv
                print('dumping data to .csv format took %f seconds'%csv_time)

            elif self.output_data_format == "excel":
                start_excel = timeit.default_timer()
                append_df_to_excel(self.output_parameters['excel_file'][0],data_to_be_saved,
                            sheet_name='Data',startrow=0)
                stop_excel = timeit.default_timer()
                excel_time = stop_excel-start_excel
                print('dumping data to .excel format took %f seconds'%excel_time)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
